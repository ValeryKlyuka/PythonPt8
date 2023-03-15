import openpyxl

class Build:
    def __init__(self, level, w, h):
        self.map = openpyxl.load_workbook(f'maps/map{level}.xlsx')
        i = 0
        code = []
        codeName = []
        for col in self.map[self.map.sheetnames[4]]:
            for row in col:
                cell = row
                if i % 2 == 0:
                    code.append(cell.fill.start_color.index)
                else:
                    codeName.append(cell.value)
                i += 1

        color_code = {}
        for i in range(len(code)):
            color_code[code[i]] = codeName[i]

        self.layer = ['', '', '', '']
        for l in range(4):
            self.layer[l] = [['1' for i in range(w)] for j in range(h)]
            i = j = 0
            for col in self.map[self.map.sheetnames[l]]:
                for row in col:
                    cell = row
                    self.layer[l][i][j] = color_code[cell.fill.start_color.index]
                    j += 1
                i += 1
                j = 0